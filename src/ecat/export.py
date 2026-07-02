"""Export helpers for processed data and figures."""

import os
import re

import numpy as np
import pandas as pd

from .utils import _format_path_for_display


def save_data(object_list, options=None):
    options = {} if options is None else dict(options)
    export_format = str(options.get("format", "csv")).strip().lower().lstrip(".")
    if export_format in {"xlsx", "excel"}:
        return _save_data_xlsx(object_list, options)
    return _save_data_csv(object_list, options)


def _save_data_csv(object_list, options=None):
    """
    Export electrochem data with two header rows:
      - Row 1: object/group 'name'
      - Row 2: column name with units in parentheses

    Units policy:
      - Source units are read ONLY from e.units.get(<column_name>).
      - If options request a target unit, values are converted using the
        object's ``scale_axis`` method or the module-level eCAT fallback.
      - If no target is requested, values are kept as-is and labeled with the
        source unit when available.
    """
    debug = bool(options.get("debug", False))
    folder_path = options.get("folder path", ".")
    file_name = options.get("file name", "output")
    output_path = os.path.join(folder_path, f"{file_name}.csv")

    if not object_list:
        raise ValueError("object_list is empty")

    def dprint(*args):
        if debug:
            print(*args)

    paren_re = re.compile(r"\s*\(.*\)\s*$")

    def with_unit(colname: str, unit: str | None) -> str:
        name = str(colname)
        if not unit or paren_re.search(name):
            return name
        return f"{name} ({unit})"

    def norm_units_dict(d):
        return {str(k): (None if v is None else str(v)) for k, v in d.items()} if isinstance(d, dict) else {}

    options_units = norm_units_dict(options.get("units", {}))
    global_x_unit = options.get("x unit", None)
    global_y_unit = options.get("y unit", None)

    def as_df(obj) -> pd.DataFrame:
        if isinstance(obj, pd.DataFrame):
            return obj.reset_index(drop=True)
        if isinstance(obj, pd.Series):
            return obj.to_frame().reset_index(drop=True)
        arr = np.asarray(obj)
        if arr.ndim == 1:
            return pd.DataFrame(arr).reset_index(drop=True)
        return pd.DataFrame(arr).reset_index(drop=True)

    def get_raw_x_series(e) -> pd.Series:
        try:
            x_raw = e.x({"one column": False})
        except TypeError:
            x_raw = e.x()
        X = as_df(x_raw)
        if X.shape[1] > 1:
            X = X.iloc[:, [0]]
        name = X.columns[0] if X.columns[0] is not None and not str(X.columns[0]).startswith("Unnamed") else "x"
        s = X.iloc[:, 0].reset_index(drop=True)
        s.name = str(name)
        return s

    def same_series(a: pd.Series, b: pd.Series, rtol=1e-10, atol=1e-12) -> bool:
        av = np.asarray(a).reshape(-1)
        bv = np.asarray(b).reshape(-1)
        if av.shape != bv.shape:
            return False
        if np.array_equal(av, bv, equal_nan=True):
            return True
        try:
            return np.allclose(av.astype(float), bv.astype(float), rtol=rtol, atol=atol, equal_nan=True)
        except Exception:
            return False

    def select_target_unit(colname: str, is_x: bool) -> str | None:
        if colname in options_units and options_units[colname] is not None:
            return options_units[colname]
        return global_x_unit if is_x else global_y_unit

    def scale_fn_for(e):
        fn = getattr(e, "scale_axis", None)
        if callable(fn):
            return fn
        from .utils import scale_axis

        return scale_axis

    def scale_and_label_series(e, s: pd.Series, is_x: bool) -> pd.DataFrame:
        col = s.name
        units_dict = getattr(e, "units", {}) or {}
        current_unit = units_dict.get(col)
        target_unit = select_target_unit(col, is_x)

        fn = scale_fn_for(e)
        if target_unit is not None:
            scale, out_unit = fn(s.values, col, current_unit, target_unit)
            out_unit = out_unit if out_unit is not None else current_unit
            values = s.astype(float) * float(scale)
            label = with_unit(col, out_unit)
        else:
            values = s
            label = with_unit(col, current_unit)

        return pd.DataFrame({label: values})

    def scale_and_label_frame(e, Y: pd.DataFrame, is_x: bool = False) -> pd.DataFrame:
        parts = []
        for c in Y.columns:
            col_series = pd.Series(Y[c].values, name=str(c))
            parts.append(scale_and_label_series(e, col_series, is_x=is_x))
        return pd.concat(parts, axis=1)

    def is_grouped(seq) -> bool:
        return isinstance(seq[0], (list, tuple))

    echem_object_count = len(object_list)
    grouped = is_grouped(object_list)
    dprint(f"Detected grouped={grouped}")

    if grouped:
        frames = []
        for g_idx, group in enumerate(object_list):
            if not group:
                continue
            echem_object_count += len(group) - 1

            raw_xs = [get_raw_x_series(e) for e in group]
            names = [getattr(e, "name", "Unnamed") for e in group]

            Ys_raw = []
            for e in group:
                Y = as_df(e.y(options))
                Y.columns = [
                    str(c) if c is not None and not str(c).startswith("Unnamed") else f"y{i + 1}"
                    for i, c in enumerate(Y.columns)
                ]
                Ys_raw.append(Y)

            max_len = max(max(len(s) for s in raw_xs), max(y.shape[0] for y in Ys_raw))
            raw_xs = [s.reindex(range(max_len)) for s in raw_xs]
            Ys_raw = [y.reindex(range(max_len)) for y in Ys_raw]

            x_all_same = all(same_series(raw_xs[0], s) for s in raw_xs[1:])
            dprint(f"[Group {g_idx}] names={names}, x_all_same={x_all_same}")

            if x_all_same:
                e0 = group[0]
                X_shared = scale_and_label_series(e0, raw_xs[0], is_x=True)
                y_blocks = [scale_and_label_frame(e, Y) for e, Y in zip(group, Ys_raw)]
                parts = [X_shared] + y_blocks
                block = pd.concat(parts, axis=1)

                top_labels = [""] + [nm for nm, Y in zip(names, y_blocks) for _ in range(Y.shape[1])]
                bottom_labels = [X_shared.columns[0]] + [c for Y in y_blocks for c in Y.columns]
                block.columns = pd.MultiIndex.from_tuples(list(zip(top_labels, bottom_labels)))
            else:
                subs = []
                for nm, e, sX, Y in zip(names, group, raw_xs, Ys_raw):
                    X_conv = scale_and_label_series(e, sX, is_x=True)
                    Y_conv = scale_and_label_frame(e, Y)
                    sub = pd.concat([X_conv, Y_conv], axis=1)
                    sub.columns = pd.MultiIndex.from_product([[nm], list(sub.columns)])
                    subs.append(sub)
                block = pd.concat(subs, axis=1)

            frames.append(block)

        all_data = pd.concat(frames, axis=1)

    else:
        blocks = []
        for i, e in enumerate(object_list):
            nm = getattr(e, "name", "Unnamed")

            if hasattr(e, "data") and isinstance(e.data, (pd.DataFrame, pd.Series)):
                df = as_df(e.data)
                df.columns = [
                    str(c) if c is not None and not str(c).startswith("Unnamed") else f"c{i + 1}"
                    for i, c in enumerate(df.columns)
                ]
                df_conv = scale_and_label_frame(e, df, is_x=False)
                df_conv.columns = pd.MultiIndex.from_product([[nm], list(df_conv.columns)])
                dprint(f"[Ungrouped {i}] using .data with shape {df.shape}")
            else:
                sX = get_raw_x_series(e)
                Y = as_df(e.y(options))
                Y.columns = [
                    str(c) if c is not None and not str(c).startswith("Unnamed") else f"y{i + 1}"
                    for i, c in enumerate(Y.columns)
                ]

                max_len = max(len(sX), Y.shape[0])
                sX = sX.reindex(range(max_len))
                Y = Y.reindex(range(max_len))

                X_conv = scale_and_label_series(e, sX, is_x=True)
                Y_conv = scale_and_label_frame(e, Y)
                df_conv = pd.concat([X_conv, Y_conv], axis=1)
                df_conv.columns = pd.MultiIndex.from_product([[nm], list(df_conv.columns)])
                dprint(f"[Ungrouped {i}] built x|y with shapes X={X_conv.shape}, Y={Y_conv.shape}")

            blocks.append(df_conv)

        all_data = pd.concat(blocks, axis=1)

    os.makedirs(folder_path, exist_ok=True)
    all_data.to_csv(output_path, index=False)
    abs_path = os.path.abspath(output_path)
    print(
        "Saved "
        f"{echem_object_count} echem objects to:\n"
        f"{_format_path_for_display(abs_path)}"
    )
    return all_data


def _flatten_echem_objects(object_list):
    flat = []
    for item in object_list:
        if isinstance(item, (list, tuple)):
            flat.extend(_flatten_echem_objects(item))
        else:
            flat.append(item)
    return flat


def _class_key(obj):
    name = type(obj).__name__.lower()
    type_text = str(getattr(obj, "type", "") or "").lower()
    if name in {"cv", "ca", "cp", "dpv"}:
        return name
    if "cyclic voltammetry" in type_text:
        return "cv"
    if "chronoamperometry" in type_text:
        return "ca"
    if "chronopotentiometry" in type_text:
        return "cp"
    if "differential pulse" in type_text:
        return "dpv"
    return name or "echem"


def _safe_sheet_name(name, existing):
    base = re.sub(r"[\[\]:*?/\\]", "_", str(name or "sheet")).strip() or "sheet"
    base = base[:31]
    candidate = base
    i = 2
    while candidate in existing:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    existing.add(candidate)
    return candidate


def _normalize_data_column_name(column):
    if isinstance(column, tuple):
        parts = [str(part).strip() for part in column if str(part).strip()]
        return parts[-1] if parts else ""
    return str(column).strip()


def _with_virtual_reference_axis(obj, data):
    if not callable(getattr(obj, "has_reference_shift", None)):
        return data
    if not obj.has_reference_shift():
        return data
    if not callable(getattr(obj, "_shifted_potential_series", None)):
        return data

    shifted = obj._shifted_potential_series()
    ref_col = _normalize_data_column_name(shifted.name)
    if ref_col in data.columns:
        return data
    if "Potential" not in data.columns:
        return data

    insert_at = list(data.columns).index("Potential") + 1
    data = data.copy()
    data.insert(insert_at, ref_col, shifted.reset_index(drop=True))
    return data


def _select_data_columns(obj, data, data_columns):
    if data_columns is None:
        data_columns = "all"

    if isinstance(data_columns, str):
        mode = data_columns.strip().lower().replace("_", " ")
        if mode in {"all", "auto"}:
            return data
        if mode in {"stored", "data", "original"}:
            stored = getattr(obj, "data", pd.DataFrame())
            stored_columns = [
                _normalize_data_column_name(col)
                for col in getattr(stored, "columns", [])
            ]
            return data.loc[:, [col for col in stored_columns if col in data.columns]]
        requested = [_normalize_data_column_name(data_columns)]
    else:
        requested = [_normalize_data_column_name(col) for col in data_columns]

    missing = [col for col in requested if col not in data.columns]
    if missing:
        available = ", ".join(str(col) for col in data.columns)
        raise ValueError(
            "Unknown data column(s) for Excel export: "
            f"{', '.join(missing)}. Available columns: {available}"
        )
    return data.loc[:, requested]


def _object_dataframe(obj, data_columns=None):
    data = getattr(obj, "data", pd.DataFrame())
    if isinstance(data, pd.Series):
        data = data.to_frame()
    elif not isinstance(data, pd.DataFrame):
        data = pd.DataFrame(np.asarray(data))
    data = data.copy().reset_index(drop=True)
    data.columns = [_normalize_data_column_name(col) for col in data.columns]
    data = _with_virtual_reference_axis(obj, data)
    data = _select_data_columns(obj, data, data_columns)
    return data


def _object_units(obj):
    return {str(k): ("" if v is None else str(v)) for k, v in (getattr(obj, "units", {}) or {}).items()}


def _data_column_unit(obj, column):
    units = _object_units(obj)
    if column in units:
        return units[column]
    stripped = re.sub(r"\s*\(([^)]*)\)\s*$", "", str(column)).strip()
    return units.get(stripped, "")


def _series_same(a, b):
    av = np.asarray(a)
    bv = np.asarray(b)
    if av.shape != bv.shape:
        return False
    try:
        return np.allclose(av.astype(float), bv.astype(float), rtol=1e-10, atol=1e-12, equal_nan=True)
    except Exception:
        return np.array_equal(av, bv)


def _xlsx_value(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value if v not in (None, ""))
    return value


def _write_dataframe(ws, df):
    for col_idx, column in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(column))
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_xlsx_value(value))


def _write_matrix(ws, matrix):
    for row_idx, row in enumerate(matrix, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_xlsx_value(value))


def _build_manifest(objects, routing, options):
    from .plotting import (
        _coerce_display_columns,
        _object_table_column_mode,
        _resolve_object_table_columns,
        build_object_table,
        pretty_table_column_label,
    )

    table_result = build_object_table(
        objects,
        {
            "columns": "all",
            "print conditions": False,
            "sig figs": options.get("sig figs", 6),
        },
    )
    display_df = table_result[0] if isinstance(table_result, tuple) else pd.DataFrame(index=range(len(objects)))
    display_df = display_df.reset_index(drop=True)

    metadata_columns = options.get("metadata columns", "used")
    mode = _object_table_column_mode(metadata_columns)
    optional_manifest_columns = {
        "reference source",
        "reference shift",
        "reference label",
        "reference mode",
    }
    requested_pretty = set()
    if mode not in {"used", "all"}:
        available = build_object_table(objects, {"columns": "available", "print conditions": False})
        for requested_column in _coerce_display_columns(metadata_columns):
            try:
                requested = _resolve_object_table_columns([requested_column], available)
            except ValueError:
                normalized = str(requested_column).strip().lower().replace("_", " ")
                if normalized not in optional_manifest_columns:
                    raise
                requested_pretty.add(pretty_table_column_label(normalized))
            else:
                requested_pretty.update(pretty_table_column_label(col) for col in requested)

    if mode == "all":
        selected = list(display_df.columns)
    else:
        selected = [
            col for col in display_df.columns
            if col in requested_pretty
            or not all(str(v).strip() == "" for v in display_df[col].fillna(""))
        ]
        selected.extend(col for col in requested_pretty if col not in selected)

    manifest = pd.DataFrame(
        {
            "object_id": [routing[id(obj)]["object_id"] for obj in objects],
            "sheet": [routing[id(obj)]["sheet"] for obj in objects],
            "class": [routing[id(obj)]["class"] for obj in objects],
            "x group": [routing[id(obj)]["x_group"] for obj in objects],
        }
    )

    for col in selected:
        if col not in manifest.columns:
            manifest[col] = display_df[col].values if col in display_df.columns else ""

    if "Name" not in manifest.columns:
        manifest["Name"] = [getattr(obj, "name", "") for obj in objects]
    if "Software" not in manifest.columns:
        manifest["Software"] = [getattr(obj, "software", "") for obj in objects]

    routing_columns = ["object_id", "sheet", "class", "x group"]
    remaining = [col for col in manifest.columns if col not in routing_columns]
    return manifest[routing_columns + remaining]


def _make_class_sheet_matrix(class_objects, routing, *, share_x_axes=True, data_columns=None):
    data_by_obj = {id(obj): _object_dataframe(obj, data_columns) for obj in class_objects}
    units_by_obj = {id(obj): _object_units(obj) for obj in class_objects}

    signatures = []
    for obj in class_objects:
        df = data_by_obj[id(obj)]
        if df.empty:
            signatures.append(None)
            continue
        x_col = df.columns[0]
        x_unit = _data_column_unit(obj, x_col)
        signatures.append((x_col, x_unit, df[x_col].to_numpy()))

    groups = []
    used = set()
    if share_x_axes:
        for i, obj in enumerate(class_objects):
            if i in used:
                continue
            signature = signatures[i]
            if signature is None:
                groups.append([i])
                used.add(i)
                continue
            group = [i]
            for j in range(i + 1, len(class_objects)):
                if j in used or signatures[j] is None:
                    continue
                same_name = signatures[j][0] == signature[0]
                same_unit = signatures[j][1] == signature[1]
                same_values = _series_same(signatures[j][2], signature[2])
                if same_name and same_unit and same_values:
                    group.append(j)
            groups.append(group)
            used.update(group)
    else:
        groups = [[i] for i in range(len(class_objects))]

    rows = [[], [], []]
    value_columns = []
    x_group_counter = 1

    def add_column(group_label, column_name, unit, values):
        rows[0].append(group_label)
        rows[1].append(column_name)
        rows[2].append(unit)
        value_columns.append(pd.Series(values).reset_index(drop=True))

    def add_blank_column():
        rows[0].append("")
        rows[1].append("")
        rows[2].append("")
        value_columns.append(pd.Series(dtype=float))

    for block_index, group in enumerate(groups):
        if block_index > 0:
            add_blank_column()

        share_block_x = share_x_axes and len(group) > 1 and signatures[group[0]] is not None
        if share_block_x:
            first_obj = class_objects[group[0]]
            first_df = data_by_obj[id(first_obj)]
            x_col = first_df.columns[0]
            x_group = f"xg_{x_group_counter:03d}"
            x_group_counter += 1
            for idx in group:
                routing[id(class_objects[idx])]["x_group"] = x_group
            add_column(x_group, x_col, _data_column_unit(first_obj, x_col), first_df[x_col])

        for idx in group:
            obj = class_objects[idx]
            df = data_by_obj[id(obj)]
            object_id = routing[id(obj)]["object_id"]
            if not share_block_x:
                routing[id(obj)]["x_group"] = object_id
                columns = list(df.columns)
            else:
                columns = list(df.columns[1:])
            for col in columns:
                add_column(object_id, col, _data_column_unit(obj, col), df[col])

    max_len = max((len(col) for col in value_columns), default=0)
    for row_idx in range(max_len):
        rows.append([
            "" if row_idx >= len(col) else col.iloc[row_idx]
            for col in value_columns
        ])
    return rows


def _save_data_xlsx(object_list, options=None):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "Excel export requires openpyxl. Install it with `python -m pip install openpyxl`."
        ) from exc

    objects = _flatten_echem_objects(object_list)
    if not objects:
        raise ValueError("object_list is empty")

    folder_path = options.get("folder path", ".")
    file_name = options.get("file name", "output")
    output_path = os.path.join(folder_path, f"{file_name}.xlsx")
    share_x_axes = bool(options.get("share x axes", True))
    data_columns = options.get("data columns", "all")

    existing_sheets = {"manifest"}
    routing = {}
    class_groups = {}
    for idx, obj in enumerate(objects, start=1):
        key = _class_key(obj)
        class_groups.setdefault(key, []).append(obj)
        routing[id(obj)] = {
            "object_id": f"obj_{idx:03d}",
            "class": key,
            "sheet": None,
            "x_group": "",
        }

    sheet_matrices = {}
    for key, group in class_groups.items():
        sheet = _safe_sheet_name(key, existing_sheets)
        for obj in group:
            routing[id(obj)]["sheet"] = sheet
        sheet_matrices[sheet] = _make_class_sheet_matrix(
            group,
            routing,
            share_x_axes=share_x_axes,
            data_columns=data_columns,
        )

    manifest = _build_manifest(objects, routing, options)

    os.makedirs(folder_path, exist_ok=True)
    workbook = Workbook()
    manifest_ws = workbook.active
    manifest_ws.title = "manifest"
    _write_dataframe(manifest_ws, manifest)
    for sheet, matrix in sheet_matrices.items():
        ws = workbook.create_sheet(sheet)
        _write_matrix(ws, matrix)
    workbook.save(output_path)

    abs_path = os.path.abspath(output_path)
    print(
        "Saved "
        f"{len(objects)} echem objects to:\n"
        f"{_format_path_for_display(abs_path)}"
    )

    exported = {"manifest": manifest}
    for sheet, matrix in sheet_matrices.items():
        exported[sheet] = pd.DataFrame(matrix)
    return exported

__all__ = ["save_data"]
